from __future__ import annotations

import json
import math
import re
from pathlib import Path
from typing import Any, Iterable, Mapping

import openpyxl
import pandas as pd


PREFERRED_CASE_HEADERS: dict[str, str] = {
    "patient_id": "patient_id",
    "gender": "gender",
    "age": "age",
    "ecog_score": "ecog_score",
    "histology_type": "histology_type",
    "tumor_location": "tumor_location",
    "ct_stage": "ct_stage",
    "cn_stage": "cn_stage",
    "clinical_stage": "clinical_stage",
    "cea_level": "cea_level",
    "mmr_status": "mmr_status",
    "chief_complaint": "chief_complaint",
    "symptom_duration": "symptom_duration",
    "family_history": "family_history",
    "family_history_details": "family_history_details",
    "biopsy_confirmed": "biopsy_confirmed",
    "biopsy_details": "biopsy_details",
    "risk_factors": "risk_factors",
}

CASE_FIELD_ALIASES: dict[str, tuple[str, ...]] = {
    "patient_id": ("patient_id", "患者ID", "病例编号", "病人ID", "受试者编号"),
    "gender": ("gender", "性别"),
    "age": ("age", "年龄", "患者年龄", "年龄（具体）"),
    "ecog_score": ("ecog_score", "ECOG评分", "ECOG"),
    "histology_type": ("histology_type", "组织学类型", "病理类型", "组织类型"),
    "tumor_location": ("tumor_location", "肿瘤部位", "原发部位"),
    "ct_stage": ("ct_stage", "cT分期", "cT分期（具体）"),
    "cn_stage": ("cn_stage", "cN分期", "cN分期（具体）"),
    "clinical_stage": ("clinical_stage", "临床分期", "具体临床分期"),
    "cea_level": ("cea_level", "CEA水平", "CEA", "基线CEA水平"),
    "mmr_status": ("mmr_status", "MMR状态"),
    "chief_complaint": ("chief_complaint", "主诉"),
    "symptom_duration": ("symptom_duration", "症状持续时间"),
    "family_history": ("family_history", "家族史"),
    "family_history_details": ("family_history_details", "家族史详情"),
    "biopsy_confirmed": ("biopsy_confirmed", "病理活检确认"),
    "biopsy_details": ("biopsy_details", "活检详情"),
    "risk_factors": ("risk_factors", "危险因素", "风险因素"),
}

FIELD_BY_ALIAS = {
    alias: field
    for field, aliases in CASE_FIELD_ALIASES.items()
    for alias in aliases
}

RISK_FACTOR_SPLIT_PATTERN = re.compile(r"[,;；，、\n]+")
BOOLEAN_TRUE_TOKENS = {"1", "true", "yes", "y", "是", "有"}
BOOLEAN_FALSE_TOKENS = {"0", "false", "no", "n", "否", "无"}
NULLABLE_TEXT_FIELDS = {
    "chief_complaint",
    "symptom_duration",
    "family_history_details",
    "biopsy_details",
}


def _is_blank(value: Any) -> bool:
    if value is None:
        return True
    if isinstance(value, float) and math.isnan(value):
        return True
    if isinstance(value, (list, tuple, set, dict)):
        return len(value) == 0
    try:
        if pd.isna(value):
            return True
    except TypeError:
        pass
    text = str(value).strip()
    return text == "" or text.lower() == "nan"


def _clean_text(value: Any) -> str:
    if _is_blank(value):
        return ""
    return str(value).strip()


def _clean_nullable_text(value: Any) -> str | None:
    text = _clean_text(value)
    return text or None


def _coerce_int(value: Any) -> int | None:
    if _is_blank(value):
        return None
    if isinstance(value, bool):
        return int(value)
    if isinstance(value, int):
        return value
    if isinstance(value, float):
        if math.isnan(value):
            return None
        if value.is_integer():
            return int(value)
        return None

    text = str(value).strip()
    if re.fullmatch(r"-?\d+", text):
        return int(text)
    if re.fullmatch(r"-?\d+\.0+", text):
        return int(float(text))
    return None


def _coerce_float(value: Any) -> float | None:
    if _is_blank(value):
        return None
    if isinstance(value, bool):
        return float(value)
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value).strip()
    try:
        return float(text)
    except ValueError:
        return None


def _coerce_bool(value: Any) -> tuple[bool | None, bool]:
    if _is_blank(value):
        return None, False
    if isinstance(value, bool):
        return value, False
    numeric = _coerce_int(value)
    if numeric in {0, 1}:
        return bool(numeric), False

    text = _clean_text(value).lower()
    if text in BOOLEAN_TRUE_TOKENS:
        return True, False
    if text in BOOLEAN_FALSE_TOKENS:
        return False, False
    return None, True


def _coerce_gender(value: Any) -> tuple[int | None, bool]:
    if _is_blank(value):
        return None, False
    numeric = _coerce_int(value)
    if numeric in {1, 2}:
        return numeric, False

    text = _clean_text(value).lower()
    if text in {"男", "male", "m"}:
        return 1, False
    if text in {"女", "female", "f"}:
        return 2, False
    return None, True


def _coerce_ecog(value: Any) -> tuple[int | None, bool]:
    if _is_blank(value):
        return None, False
    numeric = _coerce_int(value)
    if numeric is None or not 0 <= numeric <= 5:
        return None, True
    return numeric, False


def _coerce_mmr_status(value: Any) -> tuple[int | None, bool]:
    if _is_blank(value):
        return None, False
    numeric = _coerce_int(value)
    if numeric in {1, 2}:
        return numeric, False

    text = _clean_text(value).lower().replace(" ", "")
    if any(token in text for token in ("pmmr", "mss")):
        return 1, False
    if any(token in text for token in ("dmmr", "msi-h", "msih")):
        return 2, False
    return None, True


def _normalize_t_stage(value: Any) -> tuple[str, bool]:
    if _is_blank(value):
        return "", False
    text = _clean_text(value)
    normalized = re.sub(r"^(?:c)?t", "", text, flags=re.IGNORECASE)
    normalized = normalized.strip().lower()
    if re.fullmatch(r"(?:is|x|0|1|2|3|4|4a|4b)", normalized):
        return normalized, False
    return "", True


def _normalize_n_stage(value: Any) -> tuple[str, bool]:
    if _is_blank(value):
        return "", False
    text = _clean_text(value)
    normalized = re.sub(r"^(?:c)?n", "", text, flags=re.IGNORECASE)
    normalized = normalized.strip().lower()
    if re.fullmatch(r"(?:x|0|1|1a|1b|1c|2|2a|2b)", normalized):
        return normalized, False
    return "", True


def _normalize_text_field(value: Any) -> tuple[str, bool]:
    if _is_blank(value):
        return "", False
    text = _clean_text(value)
    if re.fullmatch(r"\d+(?:\.\d+)?", text):
        return "", True
    return text, False


def _coerce_risk_factors(value: Any) -> tuple[list[str], bool]:
    if _is_blank(value):
        return [], False
    if isinstance(value, list):
        return [str(item).strip() for item in value if str(item).strip()], False

    text = _clean_text(value)
    if not text:
        return [], False

    try:
        parsed = json.loads(text)
    except json.JSONDecodeError:
        parsed = None

    if isinstance(parsed, list):
        return [str(item).strip() for item in parsed if str(item).strip()], False
    if parsed is not None:
        return [], True

    return [item.strip() for item in RISK_FACTOR_SPLIT_PATTERN.split(text) if item.strip()], False


def _value_from_row(row: Mapping[str, Any], aliases: Iterable[str]) -> Any:
    for alias in aliases:
        if alias in row and not _is_blank(row[alias]):
            return row[alias]
    return None


def normalize_case_payload(data: Mapping[str, Any]) -> tuple[dict[str, Any], dict[str, Any]]:
    normalized: dict[str, Any] = {}
    extras: dict[str, Any] = {}

    for raw_key, raw_value in data.items():
        key = str(raw_key).strip()
        canonical_field = FIELD_BY_ALIAS.get(key)
        if canonical_field is None:
            extras[key] = raw_value
            continue

        if canonical_field == "patient_id":
            patient_id = _coerce_int(raw_value)
            if patient_id is None or patient_id <= 0:
                raise ValueError(f"invalid patient_id: {raw_value!r}")
            normalized[canonical_field] = patient_id
        elif canonical_field == "gender":
            gender, invalid = _coerce_gender(raw_value)
            if invalid:
                raise ValueError(f"invalid gender: {raw_value!r}")
            normalized[canonical_field] = gender
        elif canonical_field == "age":
            age = _coerce_int(raw_value)
            if age is None:
                raise ValueError(f"invalid age: {raw_value!r}")
            normalized[canonical_field] = age
        elif canonical_field == "ecog_score":
            ecog, invalid = _coerce_ecog(raw_value)
            if invalid:
                raise ValueError(f"invalid ecog_score: {raw_value!r}")
            normalized[canonical_field] = ecog
        elif canonical_field == "histology_type":
            text, invalid = _normalize_text_field(raw_value)
            if invalid:
                raise ValueError(f"invalid histology_type: {raw_value!r}")
            normalized[canonical_field] = text
        elif canonical_field == "tumor_location":
            text, invalid = _normalize_text_field(raw_value)
            if invalid:
                raise ValueError(f"invalid tumor_location: {raw_value!r}")
            normalized[canonical_field] = text
        elif canonical_field == "ct_stage":
            stage, invalid = _normalize_t_stage(raw_value)
            if invalid:
                raise ValueError(f"invalid ct_stage: {raw_value!r}")
            normalized[canonical_field] = stage
        elif canonical_field == "cn_stage":
            stage, invalid = _normalize_n_stage(raw_value)
            if invalid:
                raise ValueError(f"invalid cn_stage: {raw_value!r}")
            normalized[canonical_field] = stage
        elif canonical_field == "clinical_stage":
            normalized[canonical_field] = _clean_text(raw_value)
        elif canonical_field == "cea_level":
            cea_level = _coerce_float(raw_value)
            if cea_level is None:
                raise ValueError(f"invalid cea_level: {raw_value!r}")
            normalized[canonical_field] = cea_level
        elif canonical_field == "mmr_status":
            mmr_status, invalid = _coerce_mmr_status(raw_value)
            if invalid:
                raise ValueError(f"invalid mmr_status: {raw_value!r}")
            normalized[canonical_field] = mmr_status
        elif canonical_field in NULLABLE_TEXT_FIELDS:
            normalized[canonical_field] = _clean_nullable_text(raw_value)
        elif canonical_field in {"family_history", "biopsy_confirmed"}:
            boolean_value, invalid = _coerce_bool(raw_value)
            if invalid:
                raise ValueError(f"invalid {canonical_field}: {raw_value!r}")
            normalized[canonical_field] = boolean_value
        elif canonical_field == "risk_factors":
            risk_factors, invalid = _coerce_risk_factors(raw_value)
            if invalid:
                raise ValueError(f"invalid risk_factors: {raw_value!r}")
            normalized[canonical_field] = risk_factors

    if "patient_id" not in normalized:
        raise ValueError("patient_id is required")

    return normalized, extras


def _existing_headers(sheet: openpyxl.worksheet.worksheet.Worksheet) -> list[str]:
    if sheet.max_row < 1:
        return []
    return [
        str(sheet.cell(row=1, column=column).value).strip()
        for column in range(1, sheet.max_column + 1)
        if sheet.cell(row=1, column=column).value is not None
    ]


def _ensure_header(sheet: openpyxl.worksheet.worksheet.Worksheet, header: str) -> int:
    headers = _existing_headers(sheet)
    for index, existing in enumerate(headers, start=1):
        if existing == header:
            return index
    column = len(headers) + 1
    sheet.cell(row=1, column=column, value=header)
    return column


def _column_for_field(sheet: openpyxl.worksheet.worksheet.Worksheet, field: str) -> int:
    headers = _existing_headers(sheet)
    for alias in CASE_FIELD_ALIASES[field]:
        if alias in headers:
            return headers.index(alias) + 1
    return _ensure_header(sheet, PREFERRED_CASE_HEADERS[field])


def _serialize_excel_value(field: str, value: Any) -> Any:
    if field == "risk_factors":
        return json.dumps(value or [], ensure_ascii=False)
    return value


def upsert_case_record(excel_path: str | Path, data: Mapping[str, Any]) -> None:
    workbook_path = Path(excel_path)
    normalized, extras = normalize_case_payload(data)

    if workbook_path.exists():
        workbook = openpyxl.load_workbook(workbook_path)
        sheet = workbook.active
    else:
        workbook = openpyxl.Workbook()
        sheet = workbook.active

    if sheet.max_row < 1 or not _existing_headers(sheet):
        for field in PREFERRED_CASE_HEADERS:
            _ensure_header(sheet, PREFERRED_CASE_HEADERS[field])

    patient_col = _column_for_field(sheet, "patient_id")
    target_row: int | None = None
    patient_id = normalized["patient_id"]

    for row in range(2, sheet.max_row + 1):
        existing_patient_id = _coerce_int(sheet.cell(row=row, column=patient_col).value)
        if existing_patient_id == patient_id:
            target_row = row
            break

    if target_row is None:
        target_row = max(sheet.max_row + 1, 2)

    for field, value in normalized.items():
        column = _column_for_field(sheet, field)
        sheet.cell(row=target_row, column=column, value=_serialize_excel_value(field, value))

    for key, value in extras.items():
        column = _ensure_header(sheet, key)
        sheet.cell(row=target_row, column=column, value=value)

    workbook_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(workbook_path)


def load_case_records(excel_path: str | Path) -> list[dict[str, Any]]:
    workbook_path = Path(excel_path)
    if not workbook_path.exists():
        return []

    dataframe = pd.read_excel(workbook_path)
    records: list[dict[str, Any]] = []

    for index, row in dataframe.iterrows():
        raw_row = row.to_dict()
        patient_id = _coerce_int(_value_from_row(raw_row, CASE_FIELD_ALIASES["patient_id"]))
        if patient_id is None or patient_id <= 0:
            continue

        gender, gender_invalid = _coerce_gender(_value_from_row(raw_row, CASE_FIELD_ALIASES["gender"]))
        age = _coerce_int(_value_from_row(raw_row, CASE_FIELD_ALIASES["age"]))
        ecog_score, ecog_invalid = _coerce_ecog(_value_from_row(raw_row, CASE_FIELD_ALIASES["ecog_score"]))
        histology_type, histology_invalid = _normalize_text_field(
            _value_from_row(raw_row, CASE_FIELD_ALIASES["histology_type"])
        )
        tumor_location, tumor_location_invalid = _normalize_text_field(
            _value_from_row(raw_row, CASE_FIELD_ALIASES["tumor_location"])
        )
        ct_stage, ct_stage_invalid = _normalize_t_stage(_value_from_row(raw_row, CASE_FIELD_ALIASES["ct_stage"]))
        cn_stage, cn_stage_invalid = _normalize_n_stage(_value_from_row(raw_row, CASE_FIELD_ALIASES["cn_stage"]))
        clinical_stage = _clean_text(_value_from_row(raw_row, CASE_FIELD_ALIASES["clinical_stage"]))
        cea_level = _coerce_float(_value_from_row(raw_row, CASE_FIELD_ALIASES["cea_level"]))
        mmr_status, mmr_invalid = _coerce_mmr_status(_value_from_row(raw_row, CASE_FIELD_ALIASES["mmr_status"]))
        family_history, _ = _coerce_bool(_value_from_row(raw_row, CASE_FIELD_ALIASES["family_history"]))
        biopsy_confirmed, _ = _coerce_bool(_value_from_row(raw_row, CASE_FIELD_ALIASES["biopsy_confirmed"]))
        risk_factors, _ = _coerce_risk_factors(_value_from_row(raw_row, CASE_FIELD_ALIASES["risk_factors"]))

        invalid_fields = [
            field_name
            for field_name, invalid in (
                ("gender", gender_invalid),
                ("histology_type", histology_invalid),
                ("tumor_location", tumor_location_invalid),
                ("ct_stage", ct_stage_invalid),
                ("cn_stage", cn_stage_invalid),
                ("mmr_status", mmr_invalid),
            )
            if invalid
        ]

        if invalid_fields:
            print(
                f"[WARNING] Skipping corrupted classification row {index + 2}: "
                f"patient_id={patient_id}, invalid_fields={','.join(invalid_fields)}"
            )
            continue

        if ecog_invalid:
            ecog_score = None

        records.append(
            {
                "patient_id": patient_id,
                "gender": gender or 0,
                "age": age or 0,
                "ecog_score": ecog_score,
                "histology_type": histology_type,
                "tumor_location": tumor_location,
                "ct_stage": ct_stage,
                "cn_stage": cn_stage,
                "clinical_stage": clinical_stage,
                "cea_level": cea_level or 0.0,
                "mmr_status": mmr_status or 0,
                "chief_complaint": _clean_nullable_text(_value_from_row(raw_row, CASE_FIELD_ALIASES["chief_complaint"])),
                "symptom_duration": _clean_nullable_text(_value_from_row(raw_row, CASE_FIELD_ALIASES["symptom_duration"])),
                "family_history": family_history,
                "family_history_details": _clean_nullable_text(
                    _value_from_row(raw_row, CASE_FIELD_ALIASES["family_history_details"])
                ),
                "biopsy_confirmed": biopsy_confirmed,
                "biopsy_details": _clean_nullable_text(_value_from_row(raw_row, CASE_FIELD_ALIASES["biopsy_details"])),
                "risk_factors": risk_factors,
            }
        )

    return records


def find_case_record(excel_path: str | Path, patient_id: int) -> dict[str, Any] | None:
    workbook_path = Path(excel_path)
    if not workbook_path.exists():
        return None

    workbook = openpyxl.load_workbook(workbook_path)
    sheet = workbook.active
    headers = _existing_headers(sheet)
    if not headers:
        return None

    patient_column = _column_for_field(sheet, "patient_id")
    for row in range(2, sheet.max_row + 1):
        existing_patient_id = _coerce_int(sheet.cell(row=row, column=patient_column).value)
        if existing_patient_id == patient_id:
            return {
                header: sheet.cell(row=row, column=column).value
                for column, header in enumerate(_existing_headers(sheet), start=1)
            }
    return None
